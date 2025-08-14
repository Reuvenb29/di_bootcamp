# gradebook.py
# Creates grades.xlsx with a "Grades" sheet, styled headers, student rows,
# and an AVERAGE row for each subject column.

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# ---- Sample data (you can tweak if needed) ----
data = {
    "Joe":   {"math": 65,  "science": 78,  "english": 98,  "gym": 89},
    "Bill":  {"math": 55,  "science": 72,  "english": 87,  "gym": 95},
    "Tim":   {"math": 100, "science": 45,  "english": 75,  "gym": 92},
    "Sally": {"math": 30,  "science": 25,  "english": 45,  "gym": 100},
    "Jane":  {"math": 100, "science": 100, "english": 100, "gym": 60},
}

# Column layout
headers = ["Student", "math", "science", "english", "gym"]

# ---- Create workbook & sheet ----
wb = Workbook()
ws = wb.active
ws.title = "Grades"

# ---- Write header row ----
ws.append(headers)

# Style: bold + background color for header
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")  # blue
center = Alignment(horizontal="center", vertical="center")

for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

# ---- Write student rows ----
start_row = 2
for r, (student, grades) in enumerate(data.items(), start=start_row):
    ws.cell(row=r, column=1, value=student)                          # Student name in col A
    ws.cell(row=r, column=2, value=grades["math"])
    ws.cell(row=r, column=3, value=grades["science"])
    ws.cell(row=r, column=4, value=grades["english"])
    ws.cell(row=r, column=5, value=grades["gym"])

# ---- Add averages row (last row) ----
last_data_row = start_row + len(data) - 1
avg_row = last_data_row + 1

ws.cell(row=avg_row, column=1, value="Average").font = Font(bold=True)

# Build AVERAGE formulas for each subject column (B..E)
# A=1, B=2, C=3, D=4, E=5
for col in range(2, 6):
    col_letter = ws.cell(row=1, column=col).column_letter
    formula = f"=AVERAGE({col_letter}{start_row}:{col_letter}{last_data_row})"
    c = ws.cell(row=avg_row, column=col, value=formula)
    c.font = Font(bold=True)
    c.alignment = center

# ---- Auto-fit column widths (simple heuristic) ----
for col in range(1, 6):
    max_len = 0
    for row in range(1, avg_row + 1):
        val = ws.cell(row=row, column=col).value
        if val is None:
            continue
        max_len = max(max_len, len(str(val)))
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(10, min(max_len + 2, 30))

# ---- Save next to script ----
out_path = Path(__file__).with_name("grades.xlsx")
wb.save(out_path)
print(f"Created: {out_path}")
