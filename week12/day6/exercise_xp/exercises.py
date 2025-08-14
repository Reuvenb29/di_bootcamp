#exercise 1

# calculator_excel.py
from openpyxl import Workbook

# Create a new workbook and select active sheet
wb = Workbook()
ws = wb.active
ws.title = "Calculator"

# Add headers
ws.append(["Number 1", "Operator", "Number 2", "Result"])

# Add rows with values and formulas
ws.append([5, "+", 3, "=A2+C2"])
ws.append([8, "-", 2, "=A3-C3"])
ws.append([4, "*", 6, "=A4*C4"])
ws.append([10, "/", 2, "=A5/C5"])

# Save the workbook
wb.save("calculator.xlsx")

print("calculator.xlsx created successfully")




#exercise 2

# Night-scented unstocked.py
# Finds plants where column H says "No" and prints their names from column A.

from openpyxl import load_workbook
from pathlib import Path
import sys

# --- Config ---
EXCEL_FILENAME = "plants.xlsx"   # put the workbook next to this script
SHEET_NAME = "Sheet1"            # per the exercise
START_ROW = 2                    # A2
NAME_COL = 1                     # Column A
CHECK_COL = 8                    # Column H (A=1 → H=8)

# --- Resolve path ---
script_dir = Path(__file__).parent
excel_path = script_dir / EXCEL_FILENAME

if not excel_path.exists():
    sys.exit(f"ERROR: Could not find {excel_path}. "
             "Place your plants workbook in the same folder as this script.")

# Load workbook (data_only=True so we read values if there are formulas)
wb = load_workbook(excel_path, data_only=True)

# Get the worksheet (fallback to active if Sheet1 isn't found)
try:
    ws = wb[SHEET_NAME]
except KeyError:
    ws = wb.active
    print(f"WARNING: '{SHEET_NAME}' not found. Using active sheet: '{ws.title}'")

row = START_ROW
unstocked = []

while True:
    name_cell = ws.cell(row=row, column=NAME_COL).value

    # Stop at first blank name cell
    if name_cell is None or str(name_cell).strip() == "":
        break

    status_val = ws.cell(row=row, column=CHECK_COL).value
    status_str = (str(status_val).strip().lower() if status_val is not None else "")

    # Match "no" case-insensitively and ignoring whitespace
    if status_str == "no":
        unstocked.append(str(name_cell).strip())

    row += 1

# Output
if unstocked:
    print("Plants NOT in stock:")
    for name in unstocked:
        print(f" - {name}")
    print(f"\nTotal not in stock: {len(unstocked)}")
else:
    print("All listed plants are in stock (no 'No' found in column H).")




# exercise 3

# exercise_3.py
# Load an Excel file, filter rows where Sales > 1000, and write filtered data back into the same file.

from pathlib import Path
import sys
import shutil
import pandas as pd

# ==== CONFIG ====
EXCEL_FILENAME = "data.xlsx"   # Put this file next to the script
SOURCE_SHEET = 0                # 0 = first sheet; or use a string like "Sheet1"
OUTPUT_SHEET = "Filtered"       # The sheet we will write/replace with results
SALES_COL_NAME = "Sales"        # Column to filter on (case-insensitive match)

# ==== RESOLVE PATHS ====
here = Path(__file__).parent
excel_path = here / EXCEL_FILENAME

if not excel_path.exists():
    sys.exit(f"ERROR: Excel file not found at {excel_path}. "
             f"Place your workbook beside this script or update EXCEL_FILENAME.")

# ==== BACKUP (safety) ====
backup_path = excel_path.with_name(excel_path.stem + ".backup.xlsx")
try:
    shutil.copy2(excel_path, backup_path)
    print(f"Backup created: {backup_path}")
except Exception as e:
    print(f"WARNING: Could not create backup: {e}")

# ==== LOAD ====
try:
    df = pd.read_excel(excel_path, sheet_name=SOURCE_SHEET)
except Exception as e:
    sys.exit(f"ERROR reading Excel: {e}")

if df.empty:
    sys.exit("ERROR: The source sheet is empty.")

# ==== FIND 'Sales' COLUMN (case-insensitive, tolerant of spaces) ====
def normalize(col: str) -> str:
    return "".join(col.split()).lower()

norm_cols = {normalize(c): c for c in df.columns}
target_key = normalize(SALES_COL_NAME)

if target_key not in norm_cols:
    # Try to guess a sales-like column (e.g., 'sales_amount')
    candidates = [c for k, c in norm_cols.items() if "sales" in k]
    if not candidates:
        sys.exit(f"ERROR: Could not find a '{SALES_COL_NAME}' column. "
                 f"Columns present: {list(df.columns)}")
    sales_col = candidates[0]
    print(f"NOTE: Using detected column: '{sales_col}'")
else:
    sales_col = norm_cols[target_key]

# Ensure numeric for comparison
df[sales_col] = pd.to_numeric(df[sales_col], errors="coerce")

# ==== FILTER ====
filtered = df[df[sales_col] > 1000].copy()

print(f"Rows in source: {len(df)}")
print(f"Rows with {sales_col} > 1000: {len(filtered)}")

# ==== WRITE BACK (same file, replace/creates 'Filtered' sheet) ====
try:
    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="a",                   # append to existing workbook
        if_sheet_exists="replace"   # replace the Filtered sheet if it exists
    ) as writer:
        filtered.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)
    print(f"Filtered data written to sheet: '{OUTPUT_SHEET}' in {excel_path.name}")
except FileNotFoundError:
    # If the file didn't exist (unlikely here), write a new one
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        filtered.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)
    print(f"Workbook created and filtered data written to '{OUTPUT_SHEET}'.")
except Exception as e:
    sys.exit(f"ERROR writing Excel: {e}")




# exercise 4

# exercise_4.py
# Reads product sales data, groups by product, plots chart, writes report to Excel.

import pandas as pd
import matplotlib.pyplot as plt
from pathlib import Path
import sys

# ==== CONFIG ====
INPUT_FILE = "productSales.xlsx"
OUTPUT_FILE = "sales_report.xlsx"
PRODUCT_COL = "product"
SALES_COL = "sales"

# ==== RESOLVE PATHS ====
here = Path(__file__).parent
input_path = here / INPUT_FILE
output_path = here / OUTPUT_FILE

if not input_path.exists():
    sys.exit(f"ERROR: Input file not found: {input_path}")

# ==== READ DATA ====
try:
    df = pd.read_excel(input_path)
except Exception as e:
    sys.exit(f"ERROR reading Excel file: {e}")

if df.empty:
    sys.exit("ERROR: Input file is empty.")

# ==== CLEAN COLUMN NAMES ====
df.columns = [c.strip().lower() for c in df.columns]
if PRODUCT_COL not in df.columns or SALES_COL not in df.columns:
    sys.exit(f"ERROR: Required columns '{PRODUCT_COL}' and '{SALES_COL}' not found. "
             f"Columns present: {list(df.columns)}")

# ==== GROUP & SUM ====
grouped = df.groupby(PRODUCT_COL, as_index=False)[SALES_COL].sum()

print("\nGrouped sales data:")
print(grouped)

# ==== PLOT CHART ====
plt.figure(figsize=(8, 5))
plt.bar(grouped[PRODUCT_COL], grouped[SALES_COL], color="skyblue")
plt.xlabel("Product")
plt.ylabel("Total Sales")
plt.title("Total Sales by Product")
plt.xticks(rotation=45)
plt.tight_layout()

# Save chart as PNG (optional)
chart_path = here / "sales_chart.png"
plt.savefig(chart_path)
print(f"\nChart saved as: {chart_path}")

# ==== EXPORT TO EXCEL ====
try:
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        grouped.to_excel(writer, index=False, sheet_name="Summary")
        writer.save()
        writer.close()
    print(f"\nSales report written to: {output_path}")
except Exception as e:
    sys.exit(f"ERROR writing Excel file: {e}")
