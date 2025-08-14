# Night-scented unstocked.py
# Finds plants where column H says "No" and prints their names from column A.

from openpyxl import load_workbook
from pathlib import Path
import sys

# --- Config ---
EXCEL_FILENAME = "plants.xlsx"   # put the workbook next to this script
SHEET_NAME = "Sheet1"            # per the exercise
START_ROW = 2                    # A2
NAME_COL = 1                     # Column A
CHECK_COL = 8                    # Column H (A=1 → H=8)

# --- Resolve path ---
script_dir = Path(__file__).parent
excel_path = script_dir / EXCEL_FILENAME

if not excel_path.exists():
    sys.exit(f"ERROR: Could not find {excel_path}. "
             "Place your plants workbook in the same folder as this script.")

# Load workbook (data_only=True so we read values if there are formulas)
wb = load_workbook(excel_path, data_only=True)

# Get the worksheet (fallback to active if Sheet1 isn't found)
try:
    ws = wb[SHEET_NAME]
except KeyError:
    ws = wb.active
    print(f"WARNING: '{SHEET_NAME}' not found. Using active sheet: '{ws.title}'")

row = START_ROW
unstocked = []

while True:
    name_cell = ws.cell(row=row, column=NAME_COL).value

    # Stop at first blank name cell
    if name_cell is None or str(name_cell).strip() == "":
        break

    status_val = ws.cell(row=row, column=CHECK_COL).value
    status_str = (str(status_val).strip().lower() if status_val is not None else "")

    # Match "no" case-insensitively and ignoring whitespace
    if status_str == "no":
        unstocked.append(str(name_cell).strip())

    row += 1

# Output
if unstocked:
    print("Plants NOT in stock:")
    for name in unstocked:
        print(f" - {name}")
    print(f"\nTotal not in stock: {len(unstocked)}")
else:
    print("All listed plants are in stock (no 'No' found in column H).")
